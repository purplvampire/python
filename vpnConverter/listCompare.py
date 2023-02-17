import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import  PatternFill, Font

# Firewall config
firewall_config = 'check_config.conf'

# Excel file
workbook_file = 'check_list.xlsx'


with open(firewall_config, 'r', encoding='utf-8') as f:
    search_1 = []
    for line in f.readlines():
        line = line.strip()
        search_1.append(line)


# Search for group users.           
begin = search_1.index('config user group')

secondSearch = []
for line in search_1[begin:]:
    secondSearch.append(line)

to_the_end = []
for line in secondSearch:
    to_the_end.append(line)
    if 'end' == line:
        break


numberRegex = re.compile(r'[ABC]\d\d\d\d\d')

number = []
for line in to_the_end:
    result = numberRegex.findall(line)
    for x in result:
        number.append(x)

# 判斷重複條件
correct_number = []
duplicate_number = []
for i in range(len(number)):
    if number.count(number[i]) < 3:
        correct_number.append(number[i]) 
    else:
        duplicate_number.append(number[i])



# 比對作業



print('''

        ============ VPN設定群組比對工具 ============

        這是用來比對防火牆設定檔中群組員編與申請表原編的工具，
        若有遺漏則欄位會顯示白色，若有重複則欄位顯示橘色，
        比對一次則顯示藍色。

        執行前請先將防火牆設定檔放在此目錄下，並提供檔案名稱及副檔名

        請選擇需要比對的工作表選項

        0.總公司人員
        1.分店營業員
        2.分店後台
        3.期貨(需手動加入)

        請注意，期貨公司人員未放入活頁簿，需手動比對喔

''')

option = input('請輸入選項: ')

if option.isdigit() == True:
    option = int(option)
else:
    sys.exit()


# 載入Excel檔
wb = load_workbook(workbook_file)

sheet1_name = wb.sheetnames[option] 
print('您所選擇的工作表為:', sheet1_name) 
sheet1 = wb[sheet1_name]


# 设置欄位顏色
fill1 = PatternFill("solid", fgColor="00FF9900")
fill2 = PatternFill("solid", fgColor="00CCFFFF")

# 讀取資料
read = []
for i in range(1, sheet1.max_row + 1):
    read.append(sheet1[f'D{i}'].value)


# 比對資料
for i in range(1, len(read) + 1):
    target = sheet1.cell(row=i, column=4).value
    for query_num in correct_number:
        if target == query_num:
            sheet1.cell(row=i, column=4).fill = fill2
    for dup_num in duplicate_number:
        if target == dup_num:
            sheet1.cell(row=i, column=4).fill = fill1
            
# with open('test_list.txt', 'w', encoding='utf-8') as f:
#     f.write(str(correct_number))

# 保存文件
wb.save(workbook_file)